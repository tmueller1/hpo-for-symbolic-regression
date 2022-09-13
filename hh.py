import numpy
import random2
import operator
import math
import xlwt;
import xlrd;
from xlutils.copy import copy;
from pythonds.basic.stack import Stack
from sympy import simplify, cos, sin, sqrt,exp, symbols
from deap import algorithms, base, creator, tools, gp
from openpyxl import load_workbook
Submodel = []
Run = 0
doc = open("Text.txt",'r')
line = doc.readline()
while line :
    Submodel.append(line),
    line = doc.readline()
doc.close()
doc = open("Text.txt",'w')
doc.truncate()
doc.close()
def not_empty(s):
    return s and s.strip()

global new_Submodel
tmp = list(filter(not_empty, Submodel))
new_Submodel = tmp
len_sub = len(new_Submodel)
for i in range(0,len_sub) :
    new_Submodel[i] = new_Submodel[i].replace("\n","")
Titel = []
MAX = 1000000000
f = open("pth.txt", 'r')
ss = f.read()
f.close()
book = xlrd.open_workbook(ss)
#sheet =pd.read_excel(r'ss', sheet_name=1)
#sheet = book.get_index()
sheet = book.sheet_by_name("Sheet1")
lensheet = -1
widthsheet = -1

Num = 1
Num_sub = 0

lensheet = sheet.nrows
widthsheet = sheet.ncols
#while 1: 
#    cell = sheet.cell(Num, 1).value
#    # print(cell)
#    if cell:
#        Num = Num + 1
#    else:
#        lensheet = Num - 1
#        break
#Num = 1
#while 1:
#    cell = sheet.cell(1, Num).value
##    if cell:
 #       Num = Num + 1
 #   else:
 ##       widthsheet = Num - 1
  #      break
train_par = 1
b_list = range(1, lensheet)
random2.seed(1)
#print("lensheet")
#print(lensheet)
#print(len(b_list))
blist_webeld = (random2.sample(b_list, (int)((lensheet - 1) * train_par)))
Num_of_generation = 45
def  getin(X):

    num = 1
    Sum = 0
    ret = []
    #print("blislen")
    #print(len(blist_webeld))
    while num <= widthsheet - 1:
        Sum = 0
        while Sum <= len(blist_webeld) - 1:
            hh = blist_webeld[Sum]
            X[num][Sum] = sheet.cell(hh, num).value
            if (num == 1):
                Y[Sum] = sheet.cell(hh, 0).value
            Sum = Sum + 1
        num = num + 1
    for i in range(0,widthsheet) :
        Titel.append(sheet.cell(0,i).value)
    print(Titel)
    doc = open('NOG.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open("NOG.txt",'w')
    doc.truncate()
    doc.close()
    sstr = sstr.strip('\n')
    if(sstr == ""):
        ret.append(500)
    else :
        ret.append((int)(sstr))
    doc = open('cxbp.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open("cxbp.txt",'w')
    doc.truncate()
    doc.close()
    sstr = sstr.strip('\n')
    if (sstr == ""):
        ret.append(0.5)
    else:
        ret.append(float(sstr))
    doc = open('mutpb.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open("mutpb.txt",'w')
    doc.truncate()
    doc.close()
    sstr = sstr.strip('\n')
    if (sstr == ""):
        ret.append(0.2)
    else:
        ret.append(float(sstr))

    doc = open('num_in_gen.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open("num_in_gen.txt", 'w')
    doc.truncate()
    doc.close()
    sstr = sstr.strip('\n')
    if (sstr == ""):
        ret.append(500)
    else:
        ret.append(int(sstr))
    return ret


X = [[0] * (lensheet - 1) for i in range(widthsheet)]
Y = [0] * (lensheet - 1)


def protectedDiv(left, right):
    if(right == 0):
        return MAX
    try:
        return left / right
    except ZeroDivisionError:
        return MAX


def protectedsqrt(hh):
    try:
        if (hh < 0):
            return MAX
        else:
            return math.sqrt(hh)
    except:
        return MAX


tag = [0 for i in range(605)]
creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
creator.create("Tree", gp.PrimitiveTree, fitness=creator.FitnessMin, size=-1)

pset = gp.PrimitiveSet(name="MAIN", arity=widthsheet - 1 + len(new_Submodel))
pset.addPrimitive(operator.add, arity=2)
pset.addPrimitive(operator.sub, arity=2)
pset.addPrimitive(operator.mul, arity=2)
pset.addPrimitive(protectedDiv, arity=2)
pset.addPrimitive(protectedsqrt, arity=1)
#pset.addPrimitive(math.cos, 1)
#pset.addPrimitive(math.sin, 1)
#pset.addEphemeralConstant("rand101", lambda: 10*random2.random()) # add ephemeral constant to primitive set
# pset.addPrimitive(operator.,arity=1)


def sim(tmmp):
    x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12, x13, x14, x15, x16, x17 ,x18,x19,x20,x21,x22,x23,x24 = symbols(
        'ARG0 ARG1 ARG2 ARG3 ARG4 ARG5 ARG6 ARG7 ARG8 ARG9 ARG10 ARG11 ARG12 ARG13 ARG14 ARG15 ARG16 ARG17 ARG18 ARG19 ARG20 ARG21 ARG22 ARG23')
    f = simplify(tmmp)
    return f


def change(str, leng):
    step = 0
    ret = []
    Len = 0
    save = Stack()
    while step <= leng - 1:
        tag = 0
        if (str[step] == 'p'):
            tag = 1
            step = step + 9
            continue
        if (str[step] == 'a'):
            tag = 1
            save.push(1)
            step = step + 3
            continue
        if (str[step] == 's' and str[step + 1] == 'u'):
            tag = 1
            save.push(2)
            step = step + 3
            continue
        if (str[step] == 'm'):
            tag = 1
            save.push(3)
            step = step + 3
            continue
        if (str[step] == 'D'):
            tag = 1
            save.push(4)
            step = step + 3
            continue
        if (str[step] == ','):
            tag = 1
            c = save.pop()
            if (c == 1):
                ret.append('+')
            if (c == 2):
                ret.append('-')
            if (c == 3):
                ret.append('*')
            if (c == 4):
                ret.append('/')
            step = step + 1
            continue
        if (str[step] == 'A' and str[step+1] == 'd'):

            ret.append('+')
            step = step + 3
        if (tag == 0):
            if (str[step] != ' '):
                ret.append(str[step])
            step = step + 1
    hhhh = ''.join(ret)
    return hhhh


def calen(ss):
    step = 0
    ret = 0
    while step <= len(ss) - 1:
        if (ss[step] == '/'):
            ret = ret + 1
            step = step + 1
            continue
        if (ss[step] == '+'):
            ret = ret + 1
            step = step + 1
            continue
        if (ss[step] == '*'):
            ret = ret + 1
            step = step + 1
            continue
        if (ss[step] == '/'):
            ret = ret + 1
            step = step + 1
            continue
        if (ss[step] == 'c'):
            ret = ret + 1
            step = step + 3
            continue
        if (ss[step] == 's'):
            ret = ret + 1
            step = step + 3
            continue
        if (ss[step] == 'A'):
            ret = ret + 1
            step = step + 1
        step = step + 1
    return ret

doc = open('para.txt','r')
sss = doc.read()
doc.close()
print("****************")
print(sss)
namda2 = 0.01
if(len(sss) > 1):
    namda2 = float(sss)
#namda2 = 0.1
def rett(str):
    step = 0
    rets = ""
    print(str)
    siz = len(str)
    print(siz)
    while(step <= siz - 1) :
        if(str[step] == 'A'):
            step = step + 3
            num = 0
            while step <= siz - 1 and str[step] <= '9'and str[step] >= '0' :
                num = num * 10
                num = num + int(str[step])
                step = step + 1
            print("))))))")
            print(num)
            rets = rets + Titel[num + 1]
        else :
            rets = rets + str[step]
            step = step + 1
    return rets
def evaluateRegression(individual, Y, X, pset):
    func = gp.compile(expr=individual, pset=pset)
    sstr = ''
    doc = open('out.txt', 'w')
    print(individual, file=doc)
    doc.close()
    doc = open('out.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open('out.txt', 'w')
    doc.truncate()
    doc.close()
    result = change(sstr, len(sstr))
    luck = 0
    hhhhh = str(result)
    mul2 = calen(hhhhh)
    sqerrors = []
    for i in range(0, len(Y)):
        tmp = []
        for j in range(1, widthsheet):
            tmp.append(X[j][i])
        sum = 0
        for j in range(0, len(new_Submodel)) :
            dic  = {}
            for k in range(1,widthsheet):
                dic[Titel[k]] = X[k][i]
            #new_Submodel[j] = new_Submodel[j].replace("sin","math.sin")
            #print("now new_")
            #print(new_Submodel[j])
            #print((dic))
            tmmp = new_Submodel[j]
            for key,value in dic.items() :
                new_Submodel[j] = new_Submodel[j].replace(key,str(value*1.00000))
            ans = eval(new_Submodel[j])
            new_Submodel[j] = tmmp
            tmp.append(ans)
            sum = sum + ans
        try :
            debug = (func(*tmp) - Y[i]) ** 2
        except:
            print("debug")
            print(result)
            print(tmp)
            print("debug = ")
            print(debug)
            debug = MAX
       #print(debug)
        if("zoo" in str(debug)):
            print("zoo in debug")
            print(result)
            print(tmp)
            print(Y[i])
        sqerrors.append(debug)
    sum_of_numpy = numpy.sum(sqerrors)
    if("nan" in str(sum_of_numpy)):
        print("nan error")
        print("result ist")
        print(result)
        print("sqerror ist")
        print(sqerrors)
    if("zoo" in str(sum_of_numpy)):
        print("zoo error")
        print("result ist")
        print(result)
        print("sqerror ist")
        print(sqerrors)
    if(luck == 1):
        print("sqerror ist")
        print(sqerrors)
        print("sum")
        print(sum_of_numpy)
    try :
        if (sum_of_numpy < 0 ):
            #print(result)
            #print("ERROOOROORORORORO")
            #print(sqerrors)
            #print(sum_of_numpy)
            return (MAX,)
    except:
        if(luck == 1):
            print(sum_of_numpy)
            print(sqerrors)
            return (MAX,)
    ret = math.sqrt(numpy.sum(sqerrors) / len(Y)) + mul2 ** 2 * namda2
    if(ret <= 0.001) :
        return (0,)
    else :
        return (ret,);
    #except :
     #   print("ccccccc")
     #   return (MAX,)


toolbox = base.Toolbox()
toolbox.register("expr", gp.genFull, pset=pset, min_=1, max_=3)
toolbox.register("individual", tools.initIterate, creator.Tree, toolbox.expr)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)
toolbox.register("evaluate", evaluateRegression, Y=Y, X=X, pset=pset)
toolbox.register("mate", gp.cxOnePoint)
toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)
toolbox.register("select", tools.selTournament, tournsize=5)
#if __name__ == "__main__":

def display(ind):
	fitness_list = [x.fitness.values[0] for x in ind]
	fitness_min = numpy.min(fitness_list)

	ind_min_number = numpy.where(fitness_list == fitness_min)
	ind_min = ind[ind_min_number[0][0]]

	doc = open('out.txt', 'w')
	print(ind_min, file=doc)
	doc.close()
	doc = open('out.txt', 'r')
	sstr = doc.read()
	doc.close()
	doc = open('out.txt', 'w')
	doc.truncate()
	doc.close()
	result = change(sstr, len(sstr))
	hhhhh = str(result)
	mul2 = calen(hhhhh)
	rmse_min = fitness_min - mul2 ** 2 * namda2

	return rmse_min

cnm = getin(X)
print("cnm = ")
print(cnm)
print("1111111111111")
print(type(cnm[3]))
print(cnm[3])
pop = toolbox.population(n=int(cnm[3]))
hof = tools.HallOfFame(50)
stats = tools.Statistics(lambda ind: ind)
#stats.register("avg", lambda ind: numpy.mean(ind.fitness.values))
#stats.register("std", lambda ind: numpy.std(ind.fitness.values))
stats.register("min", display)
#stats.register("max", lambda ind: numpy.max(ind.fitness.values))
algorithms.eaSimple(pop, toolbox, cxpb=cnm[1], mutpb=cnm[2], ngen=cnm[0], stats=stats, halloffame=hof, verbose=True)
step = 0
while(step <= 45) :
    indd = hof[step]
    print("------------------------------------------------------------------")
    doc = open('out.txt', 'w')
    print(indd, file=doc)
    doc.close()
    doc = open('out.txt', 'r')
    sstr = doc.read()
    doc.close()
    doc = open('out.txt', 'w')
    doc.truncate()
    doc.close()
    result = change(sstr, len(sstr))
    print("result")
    print(result)
    print("len_Sub")
    print(len(new_Submodel))
    fk = ''
    fk = sim(result)
    print("after sim")
    print(fk)
    hhhh = str(fk)
    for i in  range (0 , len(new_Submodel) ) :
        mtp = "ARG" + str(widthsheet - 1 + i)
        print("mtp")
        print(mtp)
        hhhh = hhhh.replace(mtp,"(" + new_Submodel[i]+")")
    print("after")
    print(hhhh)
    for i in range (1,widthsheet) :
        ssstr = "ARG" + str(i-1)
        hhhh = hhhh.replace("ARG"+str(i - 1),Titel[i])
    print("after2")
    print(hhhh)


    if(not("nan" in hhhh or "zoo" in hhhh )) :
        print("out")
        print(hhhh)
        mul2 = calen(hhhh)
        doc = open('out.txt','w')
        print(hhhh,file = doc)
        doc.close()
        doc = open('out.txt','r')
        fk = doc.read()
        doc.close()
        doc = open('out.txt', 'w')
        doc.truncate()
        doc.close()
        fin_ans = fk
        doc = open('tmp.txt','w')
        print(fin_ans,file=doc)
        print(indd.fitness.values[0] - mul2 ** 2 * namda2,file = doc)
        print(indd.fitness.values[0] - mul2 ** 2 * namda2,file=doc)
        doc.close()
        break;
    else :
        step = step + 1
